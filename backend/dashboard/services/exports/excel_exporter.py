from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def aplicar_estilo_cabecalho(ws):
    fill = PatternFill("solid", fgColor="2F343A")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def aplicar_estilo_tabela(ws):
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")


def ajustar_colunas(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        adjusted_width = min(max_length + 3, 55)
        ws.column_dimensions[column_letter].width = adjusted_width


def formatar_moeda(ws, colunas):
    for coluna in colunas:
        for cell in ws[coluna][1:]:
            cell.number_format = 'R$ #,##0.00'


def formatar_percentual(ws, colunas):
    for coluna in colunas:
        for cell in ws[coluna][1:]:
            cell.number_format = '0.00"%"'


def exportar_contratos_excel(contratos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"

    headers = [
        "Nº Contrato",
        "Empresa",
        "Objeto",
        "Valor Contratual",
        "Total de Aditivos",
        "Total Reajustamento",
        "Reequilíbrio",
        "Valor Total",
        "Garantia",
        "Data Início",
        "Data Fim",
        "Status",
        "% Executado",
        "MSM a Executar",
        "CSM a Executar",
    ]

    ws.append(headers)

    for contrato in contratos:
        ws.append([
            contrato.numero_contrato,
            contrato.empresa,
            contrato.objeto,
            float(contrato.valor_contratual or 0),
            float(contrato.total_aditivos or 0),
            float(contrato.total_reajustamento or 0),
            float(contrato.reequilibrio or 0),
            float(contrato.valor_total or 0),
            contrato.garantia,
            contrato.data_inicio,
            contrato.data_fim,
            contrato.status,
            float(contrato.percentual_executado or 0),
            float(contrato.valor_msm_a_executar or 0),
            float(contrato.valor_csm_a_executar or 0),
        ])

    aplicar_estilo_cabecalho(ws)
    aplicar_estilo_tabela(ws)
    ajustar_colunas(ws)
    formatar_moeda(ws, ["D", "E", "F", "G", "H", "N", "O"])
    formatar_percentual(ws, ["M"])

    ws.freeze_panes = "A2"

    return wb


def exportar_medicoes_excel(medicoes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Medições"

    headers = [
        "Nº Medição",
        "Nº Contrato",
        "Mês/Ano",
        "Valor Medido",
        "Valor Pago",
        "Data Pagamento",
        "Valor Liquidado",
        "Valor Faturado",
        "Data Faturamento",
        "Medições a Processar",
        "Situação",
    ]

    ws.append(headers)

    for medicao in medicoes:
        ws.append([
            medicao.numero_medicao,
            medicao.numero_contrato,
            medicao.mes_ano,
            float(medicao.valor_medido or 0),
            float(medicao.valor_pago or 0),
            medicao.data_pagamento,
            float(medicao.valor_liquidado or 0),
            float(medicao.valor_faturado or 0),
            medicao.data_faturamento,
            float(medicao.valor_a_processar or 0),
            medicao.situacao,
        ])

    aplicar_estilo_cabecalho(ws)
    aplicar_estilo_tabela(ws)
    ajustar_colunas(ws)
    formatar_moeda(ws, ["D", "E", "G", "H", "J"])

    ws.freeze_panes = "A2"

    return wb